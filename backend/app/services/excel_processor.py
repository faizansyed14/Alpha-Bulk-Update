"""
Excel Processing Service
"""

import pandas as pd
import io
from typing import List, Dict, Optional, Tuple
from pathlib import Path
import openpyxl
import xlrd

from app.services.column_validator import ColumnValidator
from app.services.email_normalizer import EmailNormalizer
from app.services.phone_normalizer import PhoneNormalizer


class ExcelProcessor:
    """Process Excel files (.xlsx and .xls)"""
    
    def __init__(self):
        self.column_validator = ColumnValidator()
        self.email_normalizer = EmailNormalizer()
        self.phone_normalizer = PhoneNormalizer()
    
    def validate_file(self, file_content: bytes, filename: str) -> Tuple[bool, str, Optional[str]]:
        """
        Validate Excel file.
        
        Returns:
            Tuple of (is_valid, error_message, file_type)
        """
        if not file_content or len(file_content) == 0:
            return False, "File is empty", None
        
        filename_lower = filename.lower()
        
        # Check file extension
        if not filename_lower.endswith(('.xlsx', '.xls')):
            return False, f"Invalid file type. Expected .xlsx or .xls, got {Path(filename).suffix}", None
        
        # Try to detect file type
        file_type = None
        if filename_lower.endswith('.xlsx'):
            file_type = 'xlsx'
        elif filename_lower.endswith('.xls'):
            file_type = 'xls'
        
        # Try to open file to check if it's encrypted or corrupted
        try:
            if file_type == 'xlsx':
                # Try openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
                wb.close()
            elif file_type == 'xls':
                # Try xlrd
                wb = xlrd.open_workbook(file_contents=file_content)
        except Exception as e:
            error_msg = str(e).lower()
            if 'encrypted' in error_msg or 'password' in error_msg:
                return False, "File is password-protected. Please remove password protection and try again.", None
            elif 'corrupt' in error_msg or 'not a zip file' in error_msg:
                return False, "File appears to be corrupted. Please verify the file and try again.", None
            else:
                return False, f"Error reading file: {str(e)}", None
        
        return True, "", file_type
    
    def get_sheet_names(self, file_content: bytes, file_type: str) -> List[str]:
        """Get list of sheet names from Excel file"""
        try:
            if file_type == 'xlsx':
                wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
                sheet_names = wb.sheetnames
                wb.close()
                return sheet_names
            elif file_type == 'xls':
                wb = xlrd.open_workbook(file_contents=file_content)
                return wb.sheet_names()
        except Exception as e:
            raise ValueError(f"Error reading sheet names: {str(e)}")
    
    def read_sheet(
        self,
        file_content: bytes,
        file_type: str,
        sheet_name: Optional[str] = None
    ) -> pd.DataFrame:
        """
        Read a sheet from Excel file.
        
        Returns:
            DataFrame with raw data
        """
        try:
            if file_type == 'xlsx':
                df = pd.read_excel(
                    io.BytesIO(file_content),
                    sheet_name=sheet_name,
                    engine='openpyxl'
                )
            elif file_type == 'xls':
                df = pd.read_excel(
                    io.BytesIO(file_content),
                    sheet_name=sheet_name,
                    engine='xlrd'
                )
            else:
                raise ValueError(f"Unsupported file type: {file_type}")
            
            return df
        except Exception as e:
            raise ValueError(f"Error reading sheet '{sheet_name}': {str(e)}")
    
    def process_sheet(
        self,
        file_content: bytes,
        file_type: str,
        sheet_name: Optional[str] = None
    ) -> Tuple[pd.DataFrame, Dict[str, Optional[str]], List[str]]:
        """
        Process a single sheet.
        
        Returns:
            Tuple of (processed_dataframe, column_mapping, errors)
        """
        errors = []
        
        try:
            # Read sheet
            df = self.read_sheet(file_content, file_type, sheet_name)
            
            if df.empty:
                errors.append(f"Sheet '{sheet_name}' is empty")
                return pd.DataFrame(), {}, errors
            
            # Validate columns
            is_valid, mapping, missing_columns = self.column_validator.validate_columns(df)
            
            if not is_valid:
                errors.append(
                    f"Sheet '{sheet_name}' is missing required columns: {', '.join(missing_columns)}"
                )
                return pd.DataFrame(), mapping, errors
            
            # Extract and normalize columns
            processed_df = self.column_validator.extract_columns(df, mapping)
            
            # Normalize email and phone for matching (but keep originals for display)
            if 'Email' in processed_df.columns:
                processed_df['email_normalized'] = processed_df['Email'].apply(
                    lambda x: self.email_normalizer.normalize(x)
                )
            
            if 'Phone' in processed_df.columns:
                processed_df['phone_normalized'] = processed_df['Phone'].apply(
                    lambda x: self.phone_normalizer.normalize(x)
                )
            
            return processed_df, mapping, errors
        
        except Exception as e:
            errors.append(f"Error processing sheet '{sheet_name}': {str(e)}")
            return pd.DataFrame(), {}, errors
    
    def process_multiple_sheets(
        self,
        file_content: bytes,
        file_type: str,
        sheet_names: List[str]
    ) -> Tuple[pd.DataFrame, Dict[str, Dict[str, Optional[str]]], Dict[str, List[str]]]:
        """
        Process multiple sheets and combine them.
        
        Returns:
            Tuple of (combined_dataframe, sheet_mappings, sheet_errors)
        """
        all_dataframes = []
        sheet_mappings = {}
        sheet_errors = {}
        
        for sheet_name in sheet_names:
            df, mapping, errors = self.process_sheet(file_content, file_type, sheet_name)
            
            sheet_mappings[sheet_name] = mapping
            sheet_errors[sheet_name] = errors
            
            if not df.empty and not errors:
                all_dataframes.append(df)
        
        if not all_dataframes:
            return pd.DataFrame(), sheet_mappings, sheet_errors
        
        # Combine all dataframes
        combined_df = pd.concat(all_dataframes, ignore_index=True)
        
        # Remove duplicates within combined data (based on normalized email OR phone)
        # Keep last occurrence
        combined_df = combined_df.drop_duplicates(
            subset=['email_normalized', 'phone_normalized'],
            keep='last'
        )
        
        return combined_df, sheet_mappings, sheet_errors
    
    def dataframe_to_dict(self, df: pd.DataFrame) -> List[Dict]:
        """Convert DataFrame to list of dictionaries"""
        return df.to_dict('records')

